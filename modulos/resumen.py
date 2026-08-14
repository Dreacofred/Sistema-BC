"""
modulos/resumen.py

Módulo "Generador de Resumen" de lector.py, separado a su propio archivo.
Es el módulo más grande de los 4: maneja la lectura con IA de remitos,
la revisión manual por camión, y la exportación final a Excel.

MIGRADO A CLAUDE (agosto 2026): antes usaba Gemini con un prompt de texto
libre y parseo manual de JSON (buscando "{" y "}"). Ahora usa Claude con la
herramienta forzada "registrar_lectura_remito" (ver core/prompts_ia.py),
el mismo patrón que ya usa webhook.py con los cheques de WhatsApp. Esto
elimina la necesidad de "limpiar" números a mano, porque el campo ya viene
tipado como número desde la herramienta.

Se llama desde lector.py así:
modulo_resumen.mostrar(supabase, cliente_claude, user, NOMBRES_SUCURSALES, COLOR_ROJO)

OJO: el segundo parámetro ahora es un cliente de Anthropic (anthropic.Anthropic),
NO un cliente de Gemini como antes. lector.py tiene que crear ese cliente
aparte y pasarlo acá.
"""
import streamlit as st
import pandas as pd
import json
import io
import time
import re
import gc
import base64
import requests
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.prompts_ia import HERRAMIENTA_LECTURA_REMITO, instrucciones_lectura_remito

MODELO_CLAUDE = "claude-sonnet-5"


def convertir_a_numero(valor):
    v = str(valor).strip()
    if v.isdigit(): return int(v)
    return v


def _num(valor):
    """Convierte a float de forma segura. Con la herramienta forzada de Claude
    el campo ya debería venir como número, pero esto es una red de seguridad
    por si alguna vez llega None o un string raro."""
    if valor is None:
        return 0.0
    try:
        return float(valor)
    except (TypeError, ValueError):
        return 0.0


def _bloque_imagen_claude(bytes_imagen, tipo_mime):
    return {
        "type": "image",
        "source": {
            "type": "base64",
            "media_type": tipo_mime,
            "data": base64.b64encode(bytes_imagen).decode("utf-8"),
        },
    }


def _leer_remito_con_claude(cliente_claude, bytes_imagen, tipo_mime, id_orden):
    """
    Llama a Claude con la herramienta registrar_lectura_remito para leer un
    remito. Devuelve un diccionario con los datos extraídos, o None si falló
    después de los reintentos. Muestra avisos en pantalla igual que hacía la
    versión anterior con Gemini (st.warning / st.error).
    """
    intentos_restantes = 3
    tiempo_espera = 4

    while intentos_restantes > 0:
        try:
            respuesta = cliente_claude.messages.create(
                model=MODELO_CLAUDE,
                max_tokens=1024,
                system=instrucciones_lectura_remito(),
                tools=[HERRAMIENTA_LECTURA_REMITO],
                tool_choice={"type": "tool", "name": "registrar_lectura_remito"},
                messages=[{
                    "role": "user",
                    "content": [_bloque_imagen_claude(bytes_imagen, tipo_mime)],
                }],
            )
            for bloque in respuesta.content:
                if bloque.type == "tool_use":
                    return bloque.input

            st.warning(f"⚠️ Claude no devolvió los datos con la herramienta esperada para la orden #{id_orden}.")
            return None

        except Exception as e:
            error_msg = str(e)
            # Códigos típicos de saturación/error temporal de la API de Anthropic:
            # 429 (límite de uso), 529 (sobrecargado), 500 (error interno).
            es_error_temporal = any(
                codigo in error_msg for codigo in ["429", "529", "500", "overloaded", "rate_limit"]
            )

            if es_error_temporal:
                intentos_restantes -= 1
                if intentos_restantes > 0:
                    st.warning(f"⏳ Claude saturado. Reintentando orden #{id_orden} en {tiempo_espera} seg...")
                    time.sleep(tiempo_espera)
                    tiempo_espera *= 2
                else:
                    st.error(f"❌ Falló la orden #{id_orden} después de varios intentos. Intentá más tarde.")
            else:
                st.error(f"❌ Falla técnica inesperada en orden #{id_orden}: {error_msg}")
                return None

    return None


def _guardar_datos_ia_en_sesion(id_orden, d_ia):
    """Vuelca el resultado de _leer_remito_con_claude en st.session_state,
    con las mismas claves que usaba la versión con Gemini."""
    st.session_state[f"ia_fec_{id_orden}"] = str(d_ia.get('fecha') or '')
    st.session_state[f"ia_rs_{id_orden}"] = str(d_ia.get('razon_social') or '')
    st.session_state[f"ia_lts_{id_orden}"] = _num(d_ia.get('litros'))
    st.session_state[f"ia_imp_{id_orden}"] = _num(d_ia.get('importe'))
    st.session_state[f"ia_fac_{id_orden}"] = str(d_ia.get('comprobante') or '')
    st.session_state[f"ia_prod_{id_orden}"] = str(d_ia.get('detalle_productos') or '')
    st.session_state[f"ia_obs_{id_orden}"] = str(d_ia.get('observaciones_ia') or '')


def mostrar(supabase, cliente_claude, user, NOMBRES_SUCURSALES, COLOR_ROJO):
    if 'resumen_para_cliente' not in st.session_state: st.session_state.resumen_para_cliente = []
    if 'agregados_excel' not in st.session_state: st.session_state.agregados_excel = []

    st.title(f"📑 Generador de Resúmenes")
    
    if user['puesto'] == 'SUPER_ADMIN':
        st.markdown(f'<div class="tarjeta-pro" style="border-left: 5px solid #28a745; padding:15px;">🔓 <strong>Acceso SUPER ADMIN:</strong> Visualizando órdenes de TODAS las sucursales.</div>', unsafe_allow_html=True)
    else:
        st.markdown(f'<div class="tarjeta-pro" style="border-left: 5px solid {COLOR_ROJO}; padding:15px;">📍 <strong>Acceso Zonal:</strong> Visualizando solo órdenes de la sucursal <strong>{NOMBRES_SUCURSALES.get(user["sucursal_id"])}</strong>.</div>', unsafe_allow_html=True)

    try:
        query = supabase.table("ordenes_carga").select("*, clientes!inner(nombre, formato_especial, sucursal_madre_id)")
        if user['puesto'] != 'SUPER_ADMIN':
            query = query.eq("clientes.sucursal_madre_id", user['sucursal_id'])
            
        res_auditoria = query.eq("estado", "DESPACHADO").order("fecha_despacho", desc=True).execute()
        ordenes = res_auditoria.data
    except Exception as e:
        st.error(f"Error de base de datos: {e}")
        ordenes = []

    if not ordenes:
        st.info("👍 Todo al día. No hay remitos pendientes para tu sucursal.")
    else:
        df_audit = pd.DataFrame(ordenes)
        df_audit['Cliente'] = df_audit['clientes'].apply(lambda x: x['nombre'] if x else "DESCONOCIDO")
        df_audit['formato_especial'] = df_audit['clientes'].apply(lambda x: x.get('formato_especial', False) if x else False)
        df_audit = df_audit[(df_audit['url_foto'].notnull()) | (df_audit['motivo_sin_foto'].notnull())]

        col_k1, col_k2, col_k3 = st.columns(3)
        col_k1.metric("Clientes con pendientes", len(df_audit['Cliente'].unique()))
        col_k2.metric("Remitos a procesar", len(df_audit))
        col_k3.metric("Litros Peticionados", f"{df_audit['litros_pedidos'].sum():,.0f} L")
        st.markdown("<hr>", unsafe_allow_html=True)

        clientes_con_movimientos = df_audit['Cliente'].unique()
        
        st.markdown("### Seleccionar Cuenta Corriente")
        cliente_sel = st.selectbox("Seleccionar Cuenta Corriente", ["--- Seleccionar ---"] + list(clientes_con_movimientos), label_visibility="collapsed")
        
        if cliente_sel != "--- Seleccionar ---":
            filtro_cliente = df_audit[df_audit['Cliente'] == cliente_sel]
            
            c_head1, c_head2 = st.columns([2, 1])
            with c_head1:
                st.subheader(f"Operaciones de {cliente_sel}")
            with c_head2:
                clave_estado_ia = f"ia_procesada_{cliente_sel}"
                if st.session_state.get(clave_estado_ia, False):
                    st.button("✅ IA Procesada", disabled=True, use_container_width=True)
                else:
                    if st.button("🚀 Leer Datos con IA", use_container_width=True):
                        barra_p = st.progress(0)
                        ordenes_con_foto = filtro_cliente[filtro_cliente['url_foto'].notnull()]
                        total = len(ordenes_con_foto)

                        if total > 0:
                            for i, (_, fila) in enumerate(ordenes_con_foto.iterrows()):
                                try:
                                    res_img = requests.get(fila['url_foto'])
                                    tipo_mime = res_img.headers.get("Content-Type", "image/jpeg")
                                    if not tipo_mime.startswith("image/") and tipo_mime != "application/pdf":
                                        tipo_mime = "image/jpeg"

                                    d_ia = _leer_remito_con_claude(cliente_claude, res_img.content, tipo_mime, fila['id'])
                                    if d_ia is not None:
                                        _guardar_datos_ia_en_sesion(fila['id'], d_ia)

                                    del res_img
                                    gc.collect()

                                except Exception as e:
                                    st.error(f"❌ Falla técnica inesperada en orden #{fila['id']}: {e}")

                                barra_p.progress((i + 1) / total)
                                # Pausa chica entre pedidos para no saturar la API.
                                # PENDIENTE DE DEFINIR: ajustar este valor con el uso real,
                                # antes era 3 seg. porque Gemini se saturaba seguido.
                                time.sleep(1)

                        st.session_state[clave_estado_ia] = True
                        st.success("✅ Extracción completa.")
                        time.sleep(2)
                        st.rerun()

            st.markdown("<br>", unsafe_allow_html=True)

            for _, fila in filtro_cliente.iterrows():
                chofer_txt = fila['chofer'] if pd.notna(fila['chofer']) else "Sin chofer"
                es_especial = fila['formato_especial']
                estado_icono = "🟢 LISTO" if fila['id'] in st.session_state.agregados_excel else "🟠 PENDIENTE"
                
                patente_txt = fila['patente'] if pd.notna(fila['patente']) and str(fila['patente']).strip() != "" else "Sin Patente"
                with st.expander(f"{estado_icono} | Camión: {patente_txt} | Chofer: {chofer_txt}"):
                    c1, c2 = st.columns([1.5, 1]) 
                    with c1:
                        if pd.notna(fila['url_foto']) and str(fila['url_foto']).strip() != "":
                            st.image(fila['url_foto'], use_container_width=True)
                        else:
                            st.warning(f"⚠️ SIN FOTO: {fila.get('motivo_sin_foto', 'No se indicó motivo')}")
                        
                        st.markdown("<hr style='margin: 10px 0; border-top: 1px dashed #ccc;'>", unsafe_allow_html=True)
                        foto_nueva = st.file_uploader("**🔄 Reemplazar remito:**", type=["jpg", "png", "jpeg"], key=f"up_{fila['id']}", label_visibility="collapsed")
                        
                        if foto_nueva:
                            if st.button("💾 Guardar y Analizar Nueva Foto", key=f"btn_up_{fila['id']}", type="primary"):
                                with st.spinner("Subiendo imagen y analizando..."):
                                    try:
                                        timestamp = int(time.time())
                                        nombre_archivo = f"ADMIN_Orden{fila['id']}_{timestamp}_remito.jpg"
                                        file_bytes = foto_nueva.getvalue()
                                        tipo_mime = foto_nueva.type or "image/jpeg"

                                        supabase.storage.from_("remitos").upload(
                                            path=nombre_archivo, file=file_bytes, file_options={"content-type": tipo_mime}
                                        )

                                        url_publica = supabase.storage.from_("remitos").get_public_url(nombre_archivo)
                                        supabase.table("ordenes_carga").update({
                                            "url_foto": url_publica, "motivo_sin_foto": "Corregido por Auditoría"
                                        }).eq("id", fila['id']).execute()

                                        d_ia = _leer_remito_con_claude(cliente_claude, file_bytes, tipo_mime, fila['id'])
                                        if d_ia is not None:
                                            _guardar_datos_ia_en_sesion(fila['id'], d_ia)

                                        st.success("✅ ¡Foto actualizada y leída por Claude! Recargando...")
                                        time.sleep(1.5)
                                        st.rerun()

                                    except Exception as e:
                                        st.error(f"❌ Falla técnica inesperada: {e}")
                            
                    with c2:
                        if fila['id'] in st.session_state.agregados_excel:
                            st.success("✅ Añadido al reporte.")
                        else:
                            with st.form(key=f"form_aud_{fila['id']}"):
                                
                                detalles_bd = fila.get('detalle_combustibles')
                                if isinstance(detalles_bd, str):
                                    try: detalles_bd = json.loads(detalles_bd)
                                    except: detalles_bd = None
                                        
                                texto_pedido_cliente = ""
                                if detalles_bd and isinstance(detalles_bd, dict):
                                    texto_pedido_cliente = " | ".join([f"{k}: {v} L" if str(v).replace('.','',1).isdigit() else f"{k}: {v}" for k, v in detalles_bd.items()])
                                elif fila.get('litros_pedidos'):
                                    texto_pedido_cliente = f"Gas Oil 500 G2: {fila.get('litros_pedidos')} L"
                                else:
                                    texto_pedido_cliente = "Tanque Lleno / No especificado"
                                    
                                extras_bd = fila.get('articulos_extra')
                                if extras_bd and str(extras_bd).strip() and str(extras_bd).lower() != "nan":
                                    texto_pedido_cliente += f" 📦 Extras: {extras_bd}"
                                    
                                st.info(f"🛒 **Pedido original:**\n{texto_pedido_cliente}")

                                col_f1, col_f2 = st.columns(2)
                                fac_fecha = col_f1.text_input("Fecha", value=st.session_state.get(f"ia_fec_{fila['id']}", ""))
                                fac_rs = col_f2.text_input("Razón Social", value=st.session_state.get(f"ia_rs_{fila['id']}", ""))
                                
                                col_f3, col_f4 = st.columns(2)
                                fac_imp = col_f3.number_input("Importe ($)", value=st.session_state.get(f"ia_imp_{fila['id']}", 0.0))
                                fac_comp = col_f4.text_input("Nº Factura", value=st.session_state.get(f"ia_fac_{fila['id']}", ""))
                                
                                prod_ia_val = st.session_state.get(f"ia_prod_{fila['id']}", "")
                                obs_ia_val = st.session_state.get(f"ia_obs_{fila['id']}", "")
                                
                                if obs_ia_val:
                                    st.warning(f"⚠️ **{obs_ia_val}**")

                                fac_prod = st.text_input("Combustibles (IA)", value=prod_ia_val)
                                fac_obs = st.text_input("Observaciones (IA)", value=obs_ia_val)
                                
                                st.markdown("---")
                                efectivo_real_bd = fila.get('efectivo_entregado') if pd.notna(fila.get('efectivo_entregado')) else fila.get('efectivo_pedido', 0)
                                tiene_foto = pd.notna(fila['url_foto']) and str(fila['url_foto']).strip() != ""
                                lts_def = float(fila['litros_pedidos']) if tiene_foto else None
                                
                                nro_ord_gen, nro_ord_lts, nro_ord_efe = "", "", ""
                                
                                if es_especial:
                                    c_o1, c_o2 = st.columns(2)
                                    fac_lts = c_o1.number_input("Total Litros", value=st.session_state.get(f"ia_lts_{fila['id']}", lts_def))
                                    nro_ord_lts = c_o2.text_input("Nº Orden Litros", value=fila['nro_orden_litros_interna'] if pd.notna(fila['nro_orden_litros_interna']) else "")
                                    c_o3, c_o4 = st.columns(2)
                                    efectivo_final = c_o3.number_input("Efectivo Entregado", value=float(efectivo_real_bd))
                                    nro_ord_efe = c_o4.text_input("Nº Orden Efectivo", value=fila['nro_orden_efectivo_interna'] if pd.notna(fila['nro_orden_efectivo_interna']) else "")
                                else:
                                    c_o1, c_o2 = st.columns(2)
                                    fac_lts = c_o1.number_input("Total Litros", value=st.session_state.get(f"ia_lts_{fila['id']}", lts_def))
                                    nro_ord_gen = c_o2.text_input("Nº Orden Normal", value=fila['nro_orden_cliente'] if pd.notna(fila['nro_orden_cliente']) else "")
                                    efectivo_final = st.number_input("Efectivo Entregado", value=float(efectivo_real_bd))
                                    
                                if st.form_submit_button("✅ Guardar Fila"):
                                    st.session_state.agregados_excel.append(fila['id'])
                                    productos_a_guardar = []
                                    
                                    if "|" in fac_prod or ":" in fac_prod:
                                        partes = fac_prod.split("|")
                                        for p in partes:
                                            if ":" in p:
                                                nombre, lts_str = p.split(":", 1)
                                                lts_str = lts_str.replace(',', '.') 
                                                numeros = re.findall(r"[-+]?\d*\.\d+|\d+", lts_str)
                                                lts_indiv = float(numeros[0]) if numeros else 0.0
                                                productos_a_guardar.append({"nombre": nombre.strip(), "litros": lts_indiv})
                                            else:
                                                productos_a_guardar.append({"nombre": p.strip(), "litros": 0.0})
                                    else:
                                        productos_a_guardar.append({"nombre": fac_prod.strip(), "litros": fac_lts})
                                    
                                    producto_mayor = max(productos_a_guardar, key=lambda x: x['litros']) if productos_a_guardar else None

                                    for prod in productos_a_guardar:
                                        es_el_mayor = (prod == producto_mayor)
                                        
                                        efectivo_asignar = efectivo_final if es_el_mayor else 0.0
                                        importe_asignar = fac_imp if es_el_mayor else 0.0
                                        obs_asignar = fac_obs.strip() if es_el_mayor else ""
                                        
                                        st.session_state.resumen_para_cliente.append({
                                            "id_orden": int(fila['id']), 
                                            "Fecha": fac_fecha.strip(), "Chofer": chofer_txt.strip(), "Razón Social": fac_rs.strip(),
                                            "Litros": prod['litros'], 
                                            "Producto": prod['nombre'],
                                            "Observaciones": obs_asignar,
                                            "Nº Orden": convertir_a_numero(nro_ord_lts) if es_especial else convertir_a_numero(nro_ord_gen),
                                            "Importe": importe_asignar, "Nº Factura": fac_comp.strip(), "Entidad Pagadora": cliente_sel,
                                            "Efectivo": efectivo_asignar, "Nº Orden Efectivo": convertir_a_numero(nro_ord_efe) if es_especial else "-"
                                        })
                                    st.rerun()

    if st.session_state.resumen_para_cliente:
        st.markdown("<hr>", unsafe_allow_html=True)
        df_res = pd.DataFrame(st.session_state.resumen_para_cliente)
        
        df_res['Litros'] = pd.to_numeric(df_res['Litros'], errors='coerce').fillna(0)
        df_res['Importe'] = pd.to_numeric(df_res['Importe'], errors='coerce').fillna(0)
        df_res['Efectivo'] = pd.to_numeric(df_res['Efectivo'], errors='coerce').fillna(0)
        
        total_row = {col: "" for col in df_res.columns}
        total_row["Razón Social"] = "TOTALES:"
        total_row["Litros"] = df_res['Litros'].sum()
        total_row["Importe"] = df_res['Importe'].sum()
        total_row["Efectivo"] = df_res['Efectivo'].sum()
        
        df_export = pd.concat([df_res, pd.DataFrame([total_row])], ignore_index=True)

        st.subheader("📊 Resumen Final")
        cols_mostrar = [c for c in df_export.columns if c != "id_orden"]
        st.dataframe(df_export[cols_mostrar], use_container_width=True)
        
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as wr:
            df_export[cols_mostrar].to_excel(wr, index=False, sheet_name='Resumen_BC')
            ws = wr.sheets['Resumen_BC']
            
            fill_header = PatternFill(start_color="C8102E", end_color="C8102E", fill_type="solid")
            font_header = Font(color="FFFFFF", bold=True)
            font_total = Font(bold=True)
            fill_total = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            borde = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            for col_num, col_name in enumerate(cols_mostrar, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = fill_header
                cell.font = font_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde
                ws.column_dimensions[get_column_letter(col_num)].width = 18

            max_row = ws.max_row
            for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=ws.max_column):
                is_total_row = (row[0].row == max_row)
                for cell in row:
                    cell.border = borde
                    cell.alignment = Alignment(vertical="center")
                    if is_total_row:
                        cell.font = font_total
                        cell.fill = fill_total
                    header = ws.cell(row=1, column=cell.column).value
                    if header in ["Importe", "Efectivo"]: cell.number_format = '"$"#,##0.00'
                    elif header == "Litros": cell.number_format = '#,##0.00'

        st.markdown("<br>", unsafe_allow_html=True)
        c_ex1, c_ex2 = st.columns(2)
        c_ex1.download_button("📥 Descargar Excel Final", data=buf.getvalue(), file_name=f"Resumen_{cliente_sel}.xlsx", use_container_width=True)
        
        if c_ex2.button("🗑️ Limpiar (Sin Auditar)"):
            st.session_state.resumen_para_cliente, st.session_state.agregados_excel = [], []
            st.session_state.pop(f"ia_procesada_{cliente_sel}", None) 
            st.rerun()
            
        st.markdown("<div class='tarjeta-pro' style='margin-top: 30px; text-align:center;'>", unsafe_allow_html=True)
        confirmar_cierre = st.checkbox("Confirmo que ya descargué el Excel y deseo cerrar estas órdenes.")
        if st.button("✅ MARCAR COMO AUDITADAS", disabled=not confirmar_cierre, use_container_width=True):
            if st.session_state.agregados_excel:
                ordenes_procesadas_db = []
                for item in st.session_state.resumen_para_cliente:
                    id_actual = item.get("id_orden")
                    
                    if id_actual and id_actual not in ordenes_procesadas_db:
                        filas_este_remito = [x for x in st.session_state.resumen_para_cliente if x.get("id_orden") == id_actual]
                        litros_totales = sum(float(x["Litros"]) for x in filas_este_remito)
                        importe_total = sum(float(x["Importe"]) for x in filas_este_remito)
                        texto_productos = " | ".join([f"{x['Producto']}: {x['Litros']}L" for x in filas_este_remito])
                        todas_obs = [x["Observaciones"] for x in filas_este_remito if x["Observaciones"].strip() != ""]
                        texto_obs = " | ".join(todas_obs)
                        
                        supabase.table("ordenes_carga").update({
                            "estado": "AUDITADO", "litros_reales": litros_totales,
                            "numero_factura": item["Nº Factura"], "monto_factura": importe_total,
                            "producto_ia": texto_productos, "observaciones_ia": texto_obs
                        }).eq("id", id_actual).execute()
                        ordenes_procesadas_db.append(id_actual)
                        
            st.session_state.resumen_para_cliente, st.session_state.agregados_excel = [], []
            st.session_state.pop(f"ia_procesada_{cliente_sel}", None) 
            st.success("¡Lote cerrado exitosamente y guardado en la base de datos!")
            time.sleep(1.5)
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)
